import sqlite3
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os
import pytz
import sys
import smtplib
from email.mime.text import MIMEText
import traceback
import re
import time

def natural_sort_key(machine_no):
    """
    machine_no（例：F-1, F-10, F-2）を自然順序でソートするためのキー関数
    F-1, F-2, F-3, ..., F-10の順序になるように数値部分を整数として扱う
    """
    if not machine_no:
        return (0, 0)  # 空の場合は最初に来るように
    
    # 英字部分と数字部分を分離（例："F-10" → ("F", 10)）
    match = re.match(r'([A-Za-z-]+)(\d+)', str(machine_no))
    if match:
        prefix = match.group(1)
        number = int(match.group(2))
        return (prefix, number)
    else:
        # パターンにマッチしない場合は文字列として扱う
        return (str(machine_no), 0)

# --- メール通知用の設定 ---
# これらの設定値は、ご自身の環境に合わせて変更してください。
# パスワードを直接コードに書くことは推奨しません。
EMAIL_SENDER = "imai@araiseimitsu.onmicrosoft.com"
EMAIL_PASSWORD = "Arai267786"
EMAIL_RECEIVERS = [
    "takada@araiseimitsu.onmicrosoft.com",
    "imai@araiseimitsu.onmicrosoft.com",
    "n.kizaki@araiseimitsu.onmicrosoft.com"
]
SMTP_SERVER = "smtp.office365.com"
SMTP_PORT = 587

def send_error_email(error_info, program_name, program_path, subject_prefix="【エラー通知】"):
    """
    エラー発生時に指定されたアカウントへメールを送信する関数
    """
    try:
        subject = f"{subject_prefix}Pythonスクリプト実行中にエラーが発生しました"
        body = f"""
お疲れ様です。

Pythonスクリプトの実行中にエラーが発生しました。
下記に詳細を記載します。

---
日時: {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}
プログラム名: {program_name}
ファイルパス: {program_path}
エラー詳細:
{error_info}
---

お手数ですが、ご確認をお願いします。
"""
        msg = MIMEText(body, "plain", "utf-8")
        msg["Subject"] = subject
        msg["From"] = EMAIL_SENDER
        msg["To"] = ", ".join(EMAIL_RECEIVERS)

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, EMAIL_RECEIVERS, msg.as_string())
        print("エラー通知メールを送信しました。")

    except Exception as e:
        print(f"メール送信中にエラーが発生しました: {e}", file=sys.stderr)

def download_cleaning_instructions():
    """
    データベースからデータを取得し、Excelファイルに書き込むスクリプト。
    エラー発生時には、プログラム名とファイルパスをメッセージに含めて表示し、メール通知も行う。
    """
    
    # プログラム名とファイルパスを取得
    program_name = os.path.basename(__file__)
    program_path = os.path.abspath(__file__)
    
    # --- 設定 ---
    db_path = r'\\192.168.1.200\共有\製造課\ロボパット\python app\cleaning_instructions.db'
    excel_path = r'\\192.168.1.200\共有\生産管理課\洗浄指示書\洗浄指示書.xlsx'
    
    try:
        # --- 東京時間の今日の日付を取得 ---
        tokyo_tz = pytz.timezone('Asia/Tokyo')
        today = datetime.now(tokyo_tz).strftime('%Y-%m-%d')
        sheet_name = datetime.now(tokyo_tz).strftime('%m%d')

    except pytz.UnknownTimeZoneError:
        error_detail = f"タイムゾーン 'Asia/Tokyo' が見つかりません。pytzライブラリが正しくインストールされているか確認してください。"
        send_error_email(error_detail, program_name, program_path)
        print(f"{error_detail}\nプログラム名: {program_name}\nファイルパス: {program_path}", file=sys.stderr)
        return

    except Exception as e:
        error_detail = traceback.format_exc()
        send_error_email(error_detail, program_name, program_path)
        print(f"日付/タイムゾーン設定中にエラーが発生しました。\nプログラム名: {program_name}\nファイルパス: {program_path}\nエラー詳細: {e}", file=sys.stderr)
        return

    print(f"本日の日付 ({tokyo_tz}): {today}")
    print(f"シート名: {sheet_name}")

    # --- データベースからデータを取得 ---
    data_to_write = []
    conn = None
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        print("データベースに接続しました。")

        table_name = 'production_plan'
        query = f"SELECT cleaning_instruction, machine_no FROM {table_name} WHERE acquisition_date = ?"
        
        cursor.execute(query, (today,))
        rows = cursor.fetchall()
        
        if rows:
            # データを取得してリストに格納
            raw_data = []
            for row in rows:
                cleaning_instruction = row[0]
                machine_no = row[1]
                
                # cleaning_instructionが0だった場合は空欄として扱う
                if cleaning_instruction == 0:
                    cleaning_value = ""
                    machine_value = ""
                else:
                    cleaning_value = cleaning_instruction
                    # cleaning_instructionが空欄か0以外だったらmachine_noを取得
                    machine_value = machine_no if cleaning_instruction not in [None, "", 0] else ""
                
                raw_data.append({
                    'cleaning_instruction': cleaning_value,
                    'machine_no': machine_value,
                    'original_machine_no': machine_no  # ソート用の元のmachine_no
                })
            
            # machine_noの自然順序でソート
            raw_data.sort(key=lambda x: natural_sort_key(x['original_machine_no']))
            
            # ソート後のデータを最終リストに格納
            data_to_write = []
            for item in raw_data:
                data_to_write.append({
                    'cleaning_instruction': item['cleaning_instruction'],
                    'machine_no': item['machine_no']
                })
            
            print(f"{len(data_to_write)} 件のデータを取得しました。")
        else:
            print("本日の日付に一致するデータは見つかりませんでした。")

    except sqlite3.Error as e:
        error_detail = traceback.format_exc()
        send_error_email(error_detail, program_name, program_path)
        print(f"データベースエラーが発生しました: {e}\nプログラム名: {program_name}\nファイルパス: {program_path}\nデータベースパス: {db_path}", file=sys.stderr)
        return
    finally:
        if conn:
            conn.close()
            print("データベース接続を閉じました。")

    if not data_to_write:
        print("書き込むデータがないため、処理を終了します。")
        return

    # --- Excelファイルにデータを書き込む ---
    max_retry = 3
    retry_delay = 2  # 秒
    
    for attempt in range(max_retry):
        try:
            # ファイルが開かれているかチェック
            if os.path.exists(excel_path):
                # ファイルが書き込み可能かテスト
                try:
                    with open(excel_path, 'r+b') as test_file:
                        pass
                except PermissionError:
                    if attempt < max_retry - 1:
                        print(f"Excelファイルが他のプログラムで開かれています。{retry_delay}秒後にリトライします... (試行 {attempt + 1}/{max_retry})")
                        time.sleep(retry_delay)
                        continue
                    else:
                        error_detail = f"Excelファイル '{excel_path}' が他のプログラム（おそらくExcel）で開かれているため、書き込みできません。\nファイルを閉じてから再実行してください。"
                        send_error_email(error_detail, program_name, program_path, subject_prefix="【ファイルアクセスエラー】")
                        print(f"エラー: {error_detail}\nプログラム名: {program_name}\nファイルパス: {program_path}", file=sys.stderr)
                        return
                
                workbook = openpyxl.load_workbook(excel_path)
                print(f"既存のExcelファイルを開きました: {excel_path}")
            else:
                workbook = Workbook()
                print(f"新しいExcelファイルを作成します: {excel_path}")

            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]
                print(f"既存のシート '{sheet_name}' を削除しました。")
            
            # シート数が5以上の場合、最も古い日付のシートを削除
            while len(workbook.sheetnames) >= 5:
                # 日付形式のシート名（MMDD）を持つシートのみを対象とする
                date_sheets = []
                for ws_name in workbook.sheetnames:
                    if len(ws_name) == 4 and ws_name.isdigit():
                        date_sheets.append(ws_name)
                
                if date_sheets:
                    # 日付順でソートして最も古いシートを特定
                    date_sheets.sort()
                    oldest_sheet = date_sheets[0]
                    del workbook[oldest_sheet]
                    print(f"シート数制限により、最も古いシート '{oldest_sheet}' を削除しました。")
                else:
                    # 日付形式のシートがない場合は最初のシートを削除
                    if workbook.sheetnames:
                        oldest_sheet = workbook.sheetnames[0]
                        del workbook[oldest_sheet]
                        print(f"シート数制限により、シート '{oldest_sheet}' を削除しました。")
                    else:
                        break
                
            sheet = workbook.create_sheet(title=sheet_name)
            print(f"新しいシート '{sheet_name}' を作成しました。")

            for index, data in enumerate(data_to_write, start=2):
                # AK列（37番目）にcleaning_instructionを書き込み
                sheet.cell(row=index, column=37, value=data['cleaning_instruction'])
                # AH列（34番目）にmachine_noを書き込み（cleaning_instructionが空欄か0以外の場合）
                if data['machine_no']:
                    sheet.cell(row=index, column=34, value=data['machine_no'])

            if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
                if workbook["Sheet"].cell(row=1, column=1).value is None:
                    del workbook["Sheet"]
                    print("デフォルトの空のシート 'Sheet' を削除しました。")

            workbook.save(excel_path)
            print(f"Excelファイルにデータを保存しました: {excel_path}")
            break  # 成功したらループを抜ける

        except PermissionError as e:
            if attempt < max_retry - 1:
                print(f"ファイルアクセスエラーが発生しました。{retry_delay}秒後にリトライします... (試行 {attempt + 1}/{max_retry})")
                time.sleep(retry_delay)
                continue
            else:
                error_detail = f"ファイルアクセスエラー: {e}\nExcelファイルが他のプログラムで開かれている可能性があります。ファイルを閉じてから再実行してください。"
                send_error_email(error_detail, program_name, program_path, subject_prefix="【ファイルアクセスエラー】")
                print(f"エラー: {error_detail}\nプログラム名: {program_name}\nファイルパス: {program_path}\nExcelファイルパス: {excel_path}", file=sys.stderr)
                return
        except openpyxl.utils.exceptions.InvalidFileException as e:
            error_detail = traceback.format_exc()
            send_error_email(error_detail, program_name, program_path, subject_prefix="【致命的エラー】")
            print(f"エラー: {excel_path} は有効なExcelファイルではありません。\nプログラム名: {program_name}\nファイルパス: {program_path}", file=sys.stderr)
            return
        except Exception as e:
            if attempt < max_retry - 1:
                print(f"予期しないエラーが発生しました。{retry_delay}秒後にリトライします... (試行 {attempt + 1}/{max_retry}): {e}")
                time.sleep(retry_delay)
                continue
            else:
                error_detail = traceback.format_exc()
                send_error_email(error_detail, program_name, program_path)
                print(f"Excel処理中にエラーが発生しました: {e}\nプログラム名: {program_name}\nファイルパス: {program_path}\nExcelファイルパス: {excel_path}", file=sys.stderr)
                return

if __name__ == "__main__":
    download_cleaning_instructions()
